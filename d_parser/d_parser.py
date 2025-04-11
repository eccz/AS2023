import openpyxl
import json
from string import ascii_letters
from datetime import datetime
from config import SHEET_NAMES, TASK_TYPE_ATTR_NAME, TYPE_ATTR_NAME, SPECIALITY_ATTR_NAME


def el_sheet_processor(sheet):
    """
    Обрабатывает лист Excel, содержащий таблицы с элементами, и
    извлекает структуру данных для каждого элемента.

    Возвращает:
        dict: Словарь с данными по каждому элементу, включая:
            - IFC классы
            - уровни детализации (LOI200, LOI300, LOI400)
            - все атрибуты элемента согласно требованиям
    """
    elements = {}
    element = {}
    element_name = ''
    table_num = ''

    cnt = 1
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        if isinstance(row[0], str):
            if row[0].startswith('Таблица'):
                table_num = row[0].strip()
                element_name = row[1].strip()
                element = {table_num: {'element_name': element_name,
                                       'IFC': [],
                                       'LOI200': [],
                                       'LOI300': [],
                                       'LOI400': [],
                                       'el_attr_list': []
                                       }
                           }
            # Извлекаем классы IFC, игнорируя лишние символы и проверяя на латиницу
            if row[0].startswith('Класс IFC'):
                ifc = [i.strip() for i in row[2].replace(',', '').strip().split(' ') if
                       all(map(lambda c: c in ascii_letters, i))]
                element[table_num]['IFC'].extend(ifc)

            if 'атрибут' in row[0].lower():
                for row_2 in sheet.iter_rows(min_row=cnt + 1, max_row=sheet.max_row, min_col=1,
                                             max_col=sheet.max_column,
                                             values_only=True):

                    if not isinstance(row_2[0], str): break
                    if row_2[0].startswith('Таблица') or row_2[0].startswith('Примечание'): break

                    if row_2[1].strip() == TYPE_ATTR_NAME or row_2[1].strip() == TASK_TYPE_ATTR_NAME:
                        element[table_num][row_2[1].strip()] = row_2[2].strip()

                    if row_2[1].strip() == SPECIALITY_ATTR_NAME:
                        element[table_num][row_2[1].strip()] = row_2[2].strip()

                    # Уровень LOI200 — наличие значения в колонке 3 (по LOI200)
                    if isinstance(row_2[2], str):
                        element[table_num]['LOI200'].append(row_2[1])
                        # element[table_num]['LOI300'].append(row_2[1])
                        # element[table_num]['LOI400'].append(row_2[1])

                    # Уровень LOI300 — наличие значения в колонке 4 (по LOI300)
                    if isinstance(row_2[3], str):
                        # if not in loi200 сделать
                        element[table_num]['LOI300'].append(row_2[1]) if not row_2[1] in element[table_num]['LOI200'] else None

                    # Уровень LOI400 — наличие значения в колонке 5 (по LOI400)
                    if isinstance(row_2[4], str) and not row_2[4].startswith('-'):
                        element[table_num]['LOI400'].append(row_2[1]) if not row_2[1] in element[table_num]['LOI300'] else None

                    # используется для составления списка параметров: ['SPECIALITY', 'Специализация']
                    element[table_num]['el_attr_list'].append([row_2[1], row_2[0]])

                # if len(element[table_num]['LOI400']) < len(element[table_num]['LOI300']):
                #     element[table_num].pop('LOI400')
                elements.update(element)
        cnt += 1
    return elements


def gr_sheet_processor(sheet, el_name):
    """
    Обрабатывает лист требований к группам элементов по строительной части AEC_GR и возвращает соответствующие классы IFC,
    если их больше одного для элемента.
    Args:
        sheet: Лист Excel.
        el_name: Название элемента для поиска.
    Returns:
        list: Список списков с информацией об IFC [элемент, тип, значение]
    """
    res = []
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        if isinstance(row[2], str) and row[2].strip() == el_name:
            ifc_lst = [el_name, row[5].strip(), row[12].strip()]
            res.append(ifc_lst) if not ifc_lst in res else None

        # for WSHT2024:
        # if isinstance(row[1], str) and row[1].strip() == el_name:
        #     res.append([el_name, row[3].strip(), row[10].strip()])
    return res


def attr_sheet_processor(sheet):
    """
    Обрабатывает лист ATTRIBUTES, собирает:
    - Полный список атрибутов
    - Маппинг Model Studio <-> Атрибут
    Args:
        sheet: Лист ATTRIBUTES Excel
    Returns:
        tuple: (словарь всех атрибутов, словарь маппинга MS <-> параметры)
    """
    full_list, map_list = {}, {}
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        if isinstance(row[1], str):
            full_list[row[1].strip()] = row[2].strip()
            if len(row) > 5:
                map_list[row[2].strip()] = row[5].strip()
    return full_list, map_list


def parse(wb, to_json=False, to_term=False):
    """
    Основная функция для парсинга Excel-файла в структуру Python.

    Args:
        wb: Открытая рабочая книга Excel
        to_json (bool): Сохранять результат в JSON
        to_term (bool): Печатать результат в терминал

    Returns:
        dict: Словарь с полными данными по элементам, атрибутам и маппингам
    """
    res = dict(dict())

    # Обрабатываем листы, содержащие элементы
    for sheet in wb.sheetnames:
        if sheet in SHEET_NAMES:
            res.update(el_sheet_processor(wb[sheet]))

    # Если у элемента более одного IFC класса — читаем их из листа группировки
    for table_num, el_props in res.items():
        if len(el_props.get('IFC')) > 1:
            el_props['IFC'] = gr_sheet_processor(wb['AEC_GR'], el_props.get('element_name'))

    # Обрабатываем лист атрибутов
    try:
        res['full_attr_list'], res['ms_attr_mapping'] = attr_sheet_processor(wb['ATTRIBUTES'])
    except Exception as e:
        print(f'Не удается обработать столбец маппинга на листе с атрибутами --- {e}')

    if to_term:
        print(res)

    if to_json:
        dt = datetime.now().strftime("file_%Y-%m-%d_%H-%M-%S")
        with open(f'{dt}_output.json', 'w', encoding='utf-8') as outfile:
            json.dump(res, outfile, ensure_ascii=False, indent=4)

    return res


if __name__ == '__main__':
    workbook = openpyxl.load_workbook("../src/ADD_D_AS_2025_cleaned.xlsx")
    parse(workbook, to_term=True, to_json=True)
